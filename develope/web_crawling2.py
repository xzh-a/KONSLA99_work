from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import os
import time


# 리스트를 찾을 수 없을 경우 스크롤함
def scroll_iframe(driver, iframe_id):
    driver.switch_to.default_content()  # 최상위로 돌아가기
    WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it(iframe_id))
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")  # 스크롤 다운
    time.sleep(2)  # 추가 항목이 로드될 때까지 기다림


# 메인 실행 파일
def main():
    filepath = 'D:/code/develope/양주_신규매장_market_Db.xlsx'
    
    # 파일 로딩 또는 생성
    if os.path.exists(filepath):
        wb = load_workbook(filename=filepath)
    else:
        wb = Workbook()
        wb.active.append(["가게명", "가게주소", "가게번호"])
    sheet = wb.active
    
    # 드라이버 설정
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    
    # 웹사이트 접속
    # 주소를 변경(search/지역명)
    driver.get("https://map.naver.com/v5/search/양주")
    time.sleep(10)
    
    # 검색 실행
    search_location = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[1]/div/div/div/input')
    search_location.send_keys('시 음식점')
    search_location.send_keys(Keys.ENTER)
    # 신규업체검색을 위한 딜레이
    # 지도 확대도 수동 작업
    time.sleep(25)
    
    driver.switch_to.frame("searchIframe")
    
    # 가게 목록 순회
    i = 1
    while True:
        try:
            market_xpath = f'/html/body/div[3]/div/div[2]/div[1]/ul/li[{i}]/div[1]/a[1]/div/div/span[1]'
                            # /html/body/div[3]/div/div[2]/div[1]/ul/li[1]/div[1]/a/div/div/span[1]
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, market_xpath))).click()
            time.sleep(3)
            
            driver.switch_to.default_content()
            WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it("entryIframe"))
            
            name = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, 'Fc1rA'))).text
            loc = driver.find_element(By.CLASS_NAME, 'LDgIH').text
            phone_num = driver.find_element(By.CLASS_NAME, 'xlx7Q').text
            
            sheet.append([name, loc, phone_num])
            
            driver.switch_to.default_content()
            scroll_iframe(driver, "searchIframe")
            i += 1
        except Exception as e:
            print(f"Reached the end or encountered an error: {e}")
            break
    
    wb.save(filepath)
    driver.quit()

if __name__ == '__main__':
    main()
